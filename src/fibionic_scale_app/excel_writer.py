from __future__ import annotations

import importlib.util
import os
import sys
from dataclasses import dataclass
from pathlib import Path

from .models import ExcelSettings, FLOW_DOWN, FLOW_RIGHT
from .weight_precision import WEIGHT_NUMBER_FORMAT, quantize_weight_value

EXCEL_MODE_AUTO = "auto"
EXCEL_MODE_FILE = "file"
EXCEL_MODE_LIVE = "live"


@dataclass(slots=True)
class ExcelWriteResult:
    path: Path
    sheet_name: str
    cell: str
    value: float
    row: int
    column: str
    backend: str


class LiveExcelUnavailableError(RuntimeError):
    """Raised when the live Excel backend cannot be used."""


def normalize_column_name(column: str) -> str:
    normalized = column.strip().upper()
    if not normalized or not normalized.isalpha():
        raise ValueError("Die Excel-Spalte muss aus Buchstaben bestehen, z. B. A oder AB.")
    return normalized


def normalize_excel_mode(mode: str) -> str:
    normalized = (mode or EXCEL_MODE_AUTO).strip().lower()
    if normalized not in {EXCEL_MODE_AUTO, EXCEL_MODE_FILE, EXCEL_MODE_LIVE}:
        raise ValueError("Der Excel-Modus muss 'auto', 'file' oder 'live' sein.")
    return normalized


def normalize_scan_direction(direction: str) -> str:
    normalized = (direction or FLOW_DOWN).strip().lower()
    if normalized not in {FLOW_DOWN, FLOW_RIGHT}:
        raise ValueError("Die Excel-Richtung muss 'down' oder 'right' sein.")
    return normalized


def scan_direction_options() -> list[tuple[str, str]]:
    return [
        (FLOW_DOWN, "Oben nach unten"),
        (FLOW_RIGHT, "Links nach rechts"),
    ]


def current_platform_label() -> str:
    if sys.platform == "darwin":
        return "macOS"
    if sys.platform.startswith("win"):
        return "Windows"
    return sys.platform


def backend_label(backend: str) -> str:
    normalized = normalize_excel_mode(backend)
    if normalized == EXCEL_MODE_FILE:
        return "Datei-Writer"
    if normalized == EXCEL_MODE_LIVE:
        return "Live-Writer"
    return "Auto"


def live_backend_supported() -> bool:
    if sys.platform not in {"darwin", "win32"}:
        return False

    return importlib.util.find_spec("xlwings") is not None


def normalize_workbook_path(path_text: str) -> Path:
    if not path_text.strip():
        raise ValueError("Bitte zuerst eine Excel-Datei auswählen.")

    path = Path(path_text).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path

    if path.suffix.lower() != ".xlsx":
        raise ValueError("Bitte eine .xlsx-Datei verwenden.")

    blocked_reason = workbook_path_block_reason(path)
    if blocked_reason:
        raise ValueError(blocked_reason)

    return path


def workbook_path_block_reason(path: Path) -> str | None:
    normalized = path.expanduser()
    if _looks_like_onedrive_path(normalized):
        return (
            "OneDrive-Dateien werden im Logger nicht direkt unterstützt. "
            "Bitte eine lokale Excel-Datei außerhalb von OneDrive verwenden und später synchronisieren."
        )
    return None


def list_workbook_sheet_names(path_text: str) -> list[str]:
    path = normalize_workbook_path(path_text)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Das Paket 'openpyxl' fehlt. Bitte installiere zuerst die Projekt-Abhängigkeiten."
        ) from exc

    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _looks_like_onedrive_path(path: Path) -> bool:
    return any("onedrive" in part.lower() for part in path.parts)


def column_name_to_index(column: str) -> int:
    result = 0
    for char in normalize_column_name(column):
        result = (result * 26) + (ord(char) - 64)
    return result


def index_to_column_name(index: int) -> str:
    if index < 1:
        raise ValueError("Der Excel-Spaltenindex muss größer als 0 sein.")

    chars: list[str] = []
    while index:
        index, remainder = divmod(index - 1, 26)
        chars.append(chr(65 + remainder))
    return "".join(reversed(chars))


def build_cell_ref(column: str, row: int) -> str:
    return f"{normalize_column_name(column)}{max(1, int(row))}"


def _is_empty_excel_value(value) -> bool:
    return value in (None, "")


def _flatten_linear_range_values(values) -> list[object]:
    if isinstance(values, (list, tuple)):
        flattened: list[object] = []
        for item in values:
            if isinstance(item, (list, tuple)):
                flattened.extend(item)
            else:
                flattened.append(item)
        return flattened
    return [values]


def _first_empty_offset(values) -> int | None:
    for index, value in enumerate(values):
        if _is_empty_excel_value(value):
            return index
    return None


MAX_SCAN_STEPS = 10_000


def find_next_empty_position_with_getter(
    value_getter,
    column: str,
    start_row: int,
    direction: str,
) -> tuple[str, int]:
    normalized_column = normalize_column_name(column)
    normalized_direction = normalize_scan_direction(direction)
    row = max(1, int(start_row))

    if normalized_direction == FLOW_DOWN:
        for _ in range(MAX_SCAN_STEPS):
            if value_getter(build_cell_ref(normalized_column, row)) in (None, ""):
                return normalized_column, row
            row += 1
        raise RuntimeError(
            f"Keine leere Zelle in Spalte {normalized_column} innerhalb von {MAX_SCAN_STEPS} Zeilen gefunden."
        )

    column_index = column_name_to_index(normalized_column)
    for _ in range(MAX_SCAN_STEPS):
        current_column = index_to_column_name(column_index)
        if value_getter(build_cell_ref(current_column, row)) in (None, ""):
            return current_column, row
        column_index += 1
    raise RuntimeError(
        f"Keine leere Zelle in Zeile {row} innerhalb von {MAX_SCAN_STEPS} Spalten gefunden."
    )


class FileExcelBackend:
    key = EXCEL_MODE_FILE

    def detect_current_cell(self, settings: ExcelSettings) -> tuple[str, int]:
        workbook, worksheet, _ = self._open_workbook(settings)
        column, row = find_next_empty_position_with_getter(
            lambda cell: worksheet[cell].value,
            settings.column,
            settings.start_row,
            settings.direction,
        )
        workbook.close()
        return column, row

    def write_to_next_empty(self, settings: ExcelSettings, value: float) -> ExcelWriteResult:
        workbook, worksheet, path = self._open_workbook(settings)
        column, row = find_next_empty_position_with_getter(
            lambda cell: worksheet[cell].value,
            settings.column,
            settings.start_row,
            settings.direction,
        )
        cell = build_cell_ref(column, row)
        quantized_value = quantize_weight_value(value)
        worksheet[cell] = quantized_value
        worksheet[cell].number_format = WEIGHT_NUMBER_FORMAT
        workbook.save(path)
        workbook.close()

        return ExcelWriteResult(
            path=path,
            sheet_name=worksheet.title,
            cell=cell,
            value=quantized_value,
            row=max(1, int(row)),
            column=normalize_column_name(column),
            backend=self.key,
        )

    def write_value(self, settings: ExcelSettings, column: str, row: int, value: float) -> ExcelWriteResult:
        workbook, worksheet, path = self._open_workbook(settings)
        cell = build_cell_ref(column, row)
        quantized_value = quantize_weight_value(value)
        worksheet[cell] = quantized_value
        worksheet[cell].number_format = WEIGHT_NUMBER_FORMAT
        workbook.save(path)
        workbook.close()

        return ExcelWriteResult(
            path=path,
            sheet_name=worksheet.title,
            cell=cell,
            value=quantized_value,
            row=max(1, int(row)),
            column=normalize_column_name(column),
            backend=self.key,
        )

    def _open_workbook(self, settings: ExcelSettings):
        path = normalize_workbook_path(settings.path)

        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Das Paket 'openpyxl' fehlt. Bitte installiere zuerst die Projekt-Abhängigkeiten."
            ) from exc

        path.parent.mkdir(parents=True, exist_ok=True)
        existing = path.exists()

        if existing:
            workbook = load_workbook(path)
        else:
            workbook = Workbook()

        sheet_name = settings.sheet_name.strip() or "Messwerte"
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        elif len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet" and not existing:
            worksheet = workbook.active
            worksheet.title = sheet_name
        else:
            worksheet = workbook.create_sheet(sheet_name)

        return workbook, worksheet, path


class LiveExcelBackend:
    key = EXCEL_MODE_LIVE

    def detect_current_cell(self, settings: ExcelSettings) -> tuple[str, int]:
        _, worksheet, _ = self._open_workbook(settings)
        return self._detect_current_cell_fast(worksheet, settings)

    def write_to_next_empty(self, settings: ExcelSettings, value: float) -> ExcelWriteResult:
        workbook, worksheet, path = self._open_workbook(settings)
        column, row = self._detect_current_cell_fast(worksheet, settings)
        cell = build_cell_ref(column, row)
        quantized_value = quantize_weight_value(value)
        cell_range = worksheet.range(cell)
        cell_range.value = quantized_value
        try:
            cell_range.number_format = WEIGHT_NUMBER_FORMAT
        except Exception:
            pass
        workbook.save()

        return ExcelWriteResult(
            path=path,
            sheet_name=worksheet.name,
            cell=cell,
            value=quantized_value,
            row=max(1, int(row)),
            column=normalize_column_name(column),
            backend=self.key,
        )

    def write_value(self, settings: ExcelSettings, column: str, row: int, value: float) -> ExcelWriteResult:
        workbook, worksheet, path = self._open_workbook(settings)
        cell = build_cell_ref(column, row)
        quantized_value = quantize_weight_value(value)
        cell_range = worksheet.range(cell)
        cell_range.value = quantized_value
        try:
            cell_range.number_format = WEIGHT_NUMBER_FORMAT
        except Exception:
            pass
        workbook.save()

        return ExcelWriteResult(
            path=path,
            sheet_name=worksheet.name,
            cell=cell,
            value=quantized_value,
            row=max(1, int(row)),
            column=normalize_column_name(column),
            backend=self.key,
        )

    def _detect_current_cell_fast(self, worksheet, settings: ExcelSettings) -> tuple[str, int]:
        direction = normalize_scan_direction(settings.direction)
        column = normalize_column_name(settings.column)
        start_row = max(1, int(settings.start_row))

        if direction == FLOW_DOWN:
            return self._detect_next_empty_down(worksheet, column, start_row)
        return self._detect_next_empty_right(worksheet, column, start_row)

    def _detect_next_empty_down(self, worksheet, column: str, start_row: int) -> tuple[str, int]:
        search_limit_row = start_row + MAX_SCAN_STEPS - 1
        last_used_row = min(max(start_row, self._worksheet_last_used_row(worksheet)), search_limit_row)
        values = _flatten_linear_range_values(worksheet.range(f"{column}{start_row}:{column}{last_used_row}").value)
        empty_offset = _first_empty_offset(values)
        if empty_offset is not None:
            return column, start_row + empty_offset
        if last_used_row >= search_limit_row:
            raise RuntimeError(
                f"Keine leere Zelle in Spalte {column} innerhalb von {MAX_SCAN_STEPS} Zeilen gefunden."
            )
        return column, last_used_row + 1

    def _detect_next_empty_right(self, worksheet, column: str, row: int) -> tuple[str, int]:
        start_column_index = column_name_to_index(column)
        search_limit_column_index = start_column_index + MAX_SCAN_STEPS - 1
        last_used_column_index = min(
            max(start_column_index, self._worksheet_last_used_column_index(worksheet)),
            search_limit_column_index,
        )
        end_column = index_to_column_name(last_used_column_index)
        values = _flatten_linear_range_values(worksheet.range(f"{column}{row}:{end_column}{row}").value)
        empty_offset = _first_empty_offset(values)
        if empty_offset is not None:
            return index_to_column_name(start_column_index + empty_offset), row
        if last_used_column_index >= search_limit_column_index:
            raise RuntimeError(
                f"Keine leere Zelle in Zeile {row} innerhalb von {MAX_SCAN_STEPS} Spalten gefunden."
            )
        return index_to_column_name(last_used_column_index + 1), row

    def _open_workbook(self, settings: ExcelSettings):
        path = normalize_workbook_path(settings.path)
        self._configure_macos_onedrive_env(path)
        xw = self._import_xlwings()

        try:
            workbook = self._resolve_or_open_workbook(xw, path)
        except Exception as exc:  # pragma: no cover - depends on local Excel app
            raise LiveExcelUnavailableError(
                "Live-Modus konnte die lokale Excel-App nicht verbinden. "
                "Bitte prüfe, ob Microsoft Excel lokal installiert ist, die Datei im Desktop-Excel geöffnet werden kann "
                f"und der Python-Host Excel steuern darf. Originalfehler: {exc}"
            ) from exc

        try:  # pragma: no cover - depends on local Excel app
            workbook.app.visible = True
        except Exception:
            pass

        sheet_name = settings.sheet_name.strip() or "Messwerte"
        try:
            worksheet = workbook.sheets[sheet_name]
        except Exception:
            if len(workbook.sheets) == 1 and str(workbook.sheets[0].name).lower().startswith("sheet"):
                worksheet = workbook.sheets[0]
                worksheet.name = sheet_name
            else:
                worksheet = workbook.sheets.add(sheet_name, after=workbook.sheets[-1])

        return workbook, worksheet, path

    @staticmethod
    def _configure_macos_onedrive_env(path: Path) -> None:
        if sys.platform != "darwin":
            return

        resolved = path.resolve()
        original_chain = (path, *path.parents)
        resolved_chain = (resolved, *resolved.parents)
        for ancestor in (*original_chain, *resolved_chain):
            name = ancestor.name
            if name == "OneDrive":
                consumer_root = str(ancestor.expanduser())
                os.environ.setdefault("ONEDRIVE_CONSUMER_MAC", consumer_root)
                os.environ.setdefault("OneDriveConsumer", consumer_root)
                os.environ.setdefault("OneDrive", consumer_root)
                return
            if name.startswith("OneDrive - ") or name.startswith("OneDrive-"):
                commercial_root = LiveExcelBackend._preferred_macos_onedrive_root(ancestor)
                os.environ.setdefault("ONEDRIVE_COMMERCIAL_MAC", commercial_root)
                os.environ.setdefault("OneDriveCommercial", commercial_root)
                os.environ.setdefault("OneDrive", commercial_root)
                return

    @staticmethod
    def _preferred_macos_onedrive_root(root: Path) -> str:
        expanded = root.expanduser()
        resolved_root = expanded.resolve()
        home = Path.home()

        for candidate in home.glob("OneDrive*"):
            try:
                if candidate.resolve() == resolved_root:
                    return str(candidate)
            except OSError:
                continue

        return str(expanded)

    def _resolve_or_open_workbook(self, xw, path: Path):
        resolved_target = path.resolve()

        for app in self._iter_apps(xw):
            for book in app.books:
                if self._book_matches_path(book, resolved_target):
                    return book

        if path.exists():
            app = self._pick_or_create_app(xw)
            return app.books.open(str(path))

        path.parent.mkdir(parents=True, exist_ok=True)
        app = self._pick_or_create_app(xw)
        workbook = app.books.add()
        workbook.save(str(path))
        return workbook

    @staticmethod
    def _iter_apps(xw):
        try:
            return list(xw.apps)
        except Exception:  # pragma: no cover - depends on local Excel app
            return []

    @staticmethod
    def _pick_or_create_app(xw):
        apps = LiveExcelBackend._iter_apps(xw)
        if apps:
            return apps[0]
        return xw.App(visible=True, add_book=False)

    @staticmethod
    def _book_matches_path(book, target: Path) -> bool:
        fullname = getattr(book, "fullname", "") or ""
        if not fullname:
            return False

        try:
            candidate = Path(fullname).expanduser().resolve()
        except OSError:
            return False

        return candidate == target

    @staticmethod
    def _worksheet_last_used_row(worksheet) -> int:
        try:
            return max(1, int(worksheet.used_range.last_cell.row or 1))
        except Exception:
            return 1

    @staticmethod
    def _worksheet_last_used_column_index(worksheet) -> int:
        try:
            return max(1, int(worksheet.used_range.last_cell.column or 1))
        except Exception:
            return 1

    @staticmethod
    def _import_xlwings():
        if sys.platform not in {"darwin", "win32"}:
            raise LiveExcelUnavailableError("Live-Modus wird nur auf macOS und Windows unterstützt.")

        try:
            import xlwings as xw
        except ImportError as exc:
            raise LiveExcelUnavailableError(
                "Das Paket 'xlwings' fehlt. Bitte installiere zuerst die Projekt-Abhängigkeiten."
            ) from exc

        return xw


FILE_BACKEND = FileExcelBackend()
LIVE_BACKEND = LiveExcelBackend()


class ExcelSession:
    def __init__(self, settings: ExcelSettings):
        self.settings = settings
        self.active_backend = normalize_excel_mode(settings.mode)

    def update_settings(self, settings: ExcelSettings) -> None:
        self.settings = settings
        self.active_backend = normalize_excel_mode(settings.mode)

    def preview_cell(self) -> str:
        return build_cell_ref(self.settings.column, self.settings.start_row)

    def backend_display_name(self) -> str:
        return backend_label(self.active_backend)

    def detect_current_cell(self) -> tuple[str, int]:
        return self._run_backend("detect_current_cell")

    def write_value(self, value: float) -> ExcelWriteResult:
        return self._run_backend("write_to_next_empty", value)

    def _run_backend(self, method_name: str, *args):
        mode = normalize_excel_mode(self.settings.mode)
        if mode == EXCEL_MODE_FILE:
            backends = [FILE_BACKEND]
        elif mode == EXCEL_MODE_LIVE:
            backends = [LIVE_BACKEND]
        else:
            backends = [LIVE_BACKEND, FILE_BACKEND]

        last_error: Exception | None = None
        for backend in backends:
            try:
                result = getattr(backend, method_name)(self.settings, *args)
                self.active_backend = backend.key
                return result
            except LiveExcelUnavailableError as exc:
                last_error = exc
                if mode == EXCEL_MODE_AUTO:
                    continue
                raise

        if last_error is not None:
            raise last_error

        raise RuntimeError("Kein Excel-Backend verfügbar.")
