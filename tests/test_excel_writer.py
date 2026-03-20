from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from fibionic_scale_app.excel_writer import (
    EXCEL_MODE_AUTO,
    EXCEL_MODE_FILE,
    LIVE_BACKEND,
    ExcelSession,
    LiveExcelUnavailableError,
    column_name_to_index,
    list_workbook_sheet_names,
    workbook_path_block_reason,
)
from fibionic_scale_app.models import ExcelSettings, FLOW_RIGHT
from fibionic_scale_app.weight_precision import WEIGHT_NUMBER_FORMAT


class ExcelSessionTests(unittest.TestCase):
    class _DummyLastCell:
        def __init__(self, row: int, column: int):
            self.row = row
            self.column = column

    class _DummyUsedRange:
        def __init__(self, row: int, column: int):
            self.last_cell = ExcelSessionTests._DummyLastCell(row, column)

    class _DummyRange:
        def __init__(self, value):
            self.value = value

    class _DummyWorksheet:
        def __init__(self, *, column_values=None, row_values=None, used_row: int = 1, used_column: int = 1):
            self.column_values = dict(column_values or {})
            self.row_values = dict(row_values or {})
            self.used_range = ExcelSessionTests._DummyUsedRange(used_row, used_column)
            self.range_calls: list[str] = []

        def range(self, ref: str):
            self.range_calls.append(ref)
            start, end = ref.split(":")
            if start[0].isalpha() and end[0].isalpha() and start[1:].isdigit() and end[1:].isdigit():
                start_column = "".join(char for char in start if char.isalpha())
                end_column = "".join(char for char in end if char.isalpha())
                start_row = int("".join(char for char in start if char.isdigit()))
                end_row = int("".join(char for char in end if char.isdigit()))

                if start_column == end_column:
                    values = [self.column_values.get(row) for row in range(start_row, end_row + 1)]
                    if len(values) == 1:
                        return ExcelSessionTests._DummyRange(values[0])
                    return ExcelSessionTests._DummyRange(values)

                if start_row == end_row:
                    values = [
                        self.row_values.get(column_index)
                        for column_index in range(
                            column_name_to_index(start_column),
                            column_name_to_index(end_column) + 1,
                        )
                    ]
                    if len(values) == 1:
                        return ExcelSessionTests._DummyRange(values[0])
                    return ExcelSessionTests._DummyRange(values)

            raise AssertionError(f"Unerwartete Range-Abfrage: {ref}")

    def test_writes_value_to_first_empty_row_and_detects_next_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "messwerte.xlsx"
            session = ExcelSession(
                ExcelSettings(
                    path=str(path),
                    sheet_name="Produktion",
                    column="F",
                    start_row=3,
                    mode=EXCEL_MODE_FILE,
                )
            )

            result = session.write_value(12.34)
            next_column, next_row = session.detect_current_cell()

            workbook = load_workbook(path)
            worksheet = workbook["Produktion"]
            self.assertEqual(worksheet["F3"].value, 12.34)
            self.assertEqual(worksheet["F3"].number_format, WEIGHT_NUMBER_FORMAT)
            self.assertEqual(result.cell, "F3")
            self.assertEqual(next_column, "F")
            self.assertEqual(next_row, 4)
            workbook.close()

    def test_rounds_written_values_to_two_decimals(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "messwerte.xlsx"
            session = ExcelSession(
                ExcelSettings(
                    path=str(path),
                    sheet_name="Produktion",
                    column="F",
                    start_row=3,
                    mode=EXCEL_MODE_FILE,
                )
            )

            result = session.write_value(13.345)

            workbook = load_workbook(path)
            worksheet = workbook["Produktion"]
            self.assertEqual(worksheet["F3"].value, 13.35)
            self.assertEqual(worksheet["F3"].number_format, WEIGHT_NUMBER_FORMAT)
            self.assertEqual(result.value, 13.35)
            workbook.close()

    def test_detects_first_empty_column_for_horizontal_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "messwerte.xlsx"
            session = ExcelSession(
                ExcelSettings(
                    path=str(path),
                    sheet_name="Produktion",
                    column="B",
                    start_row=4,
                    direction=FLOW_RIGHT,
                    mode=EXCEL_MODE_FILE,
                )
            )

            first = session.write_value(1.23)
            second = session.write_value(4.56)
            next_column, next_row = session.detect_current_cell()

            workbook = load_workbook(path)
            worksheet = workbook["Produktion"]
            self.assertEqual(worksheet["B4"].value, 1.23)
            self.assertEqual(worksheet["C4"].value, 4.56)
            self.assertEqual(first.cell, "B4")
            self.assertEqual(second.cell, "C4")
            self.assertEqual(next_column, "D")
            self.assertEqual(next_row, 4)
            workbook.close()

    def test_auto_mode_falls_back_to_file_writer(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "messwerte.xlsx"
            session = ExcelSession(
                ExcelSettings(
                    path=str(path),
                    sheet_name="Produktion",
                    column="C",
                    start_row=2,
                    mode=EXCEL_MODE_AUTO,
                )
            )

            with patch.object(
                LIVE_BACKEND,
                "detect_current_cell",
                side_effect=LiveExcelUnavailableError("Excel nicht erreichbar"),
            ):
                with patch.object(
                    LIVE_BACKEND,
                    "write_value",
                    side_effect=LiveExcelUnavailableError("Excel nicht erreichbar"),
                ):
                    result = session.write_value(7.89)

            workbook = load_workbook(path)
            worksheet = workbook["Produktion"]
            self.assertEqual(worksheet["C2"].value, 7.89)
            self.assertEqual(result.backend, EXCEL_MODE_FILE)
            workbook.close()

    def test_lists_sheet_names_from_existing_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "messwerte.xlsx"
            session = ExcelSession(
                ExcelSettings(
                    path=str(path),
                    sheet_name="Produktion",
                    column="A",
                    start_row=1,
                    mode=EXCEL_MODE_FILE,
                )
            )
            session.write_value(1.0)

            workbook = load_workbook(path)
            workbook.create_sheet("Archiv")
            workbook.save(path)
            workbook.close()

            self.assertEqual(list_workbook_sheet_names(str(path)), ["Produktion", "Archiv"])

    def test_live_backend_path_match_is_exact(self) -> None:
        class DummyBook:
            def __init__(self, fullname: str):
                self.fullname = fullname

        from fibionic_scale_app.excel_writer import LiveExcelBackend

        with tempfile.TemporaryDirectory() as tmpdir:
            target = Path(tmpdir) / "messwerte.xlsx"
            target.touch()
            book = DummyBook(str(target))

            self.assertTrue(LiveExcelBackend._book_matches_path(book, target.resolve()))

    def test_macos_onedrive_root_is_derived_from_selected_file(self) -> None:
        from fibionic_scale_app.excel_writer import LiveExcelBackend

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir) / "Library" / "CloudStorage" / "OneDrive-fibionicGmbH"
            file_path = root / "Ordner" / "messwerte.xlsx"
            file_path.parent.mkdir(parents=True, exist_ok=True)
            file_path.touch()

            with patch("fibionic_scale_app.excel_writer.sys.platform", "darwin"):
                with patch.dict(os.environ, {}, clear=True):
                    LiveExcelBackend._configure_macos_onedrive_env(file_path)
                    configured = os.environ.get("ONEDRIVE_COMMERCIAL_MAC")
                    self.assertIsNotNone(configured)
                    self.assertEqual(Path(configured).resolve(), root.resolve())
                    self.assertEqual(os.environ.get("OneDriveCommercial"), configured)

    def test_macos_onedrive_prefers_home_alias_over_cloudstorage_root(self) -> None:
        from fibionic_scale_app.excel_writer import LiveExcelBackend

        with tempfile.TemporaryDirectory() as tmpdir:
            fake_home = Path(tmpdir) / "home"
            fake_home.mkdir()
            cloud_root = fake_home / "Library" / "CloudStorage" / "OneDrive-fibionicGmbH"
            cloud_root.mkdir(parents=True)
            alias = fake_home / "OneDrive - FIBIONIC"
            alias.symlink_to(cloud_root)

            with patch("fibionic_scale_app.excel_writer.Path.home", return_value=fake_home):
                preferred = LiveExcelBackend._preferred_macos_onedrive_root(cloud_root)

            self.assertEqual(preferred, str(alias))

    def test_onedrive_path_is_blocked_with_clear_message(self) -> None:
        path = Path("C:/Users/test/OneDrive - fibionic/messwerte.xlsx")

        reason = workbook_path_block_reason(path)

        self.assertIsNotNone(reason)
        self.assertIn("OneDrive-Dateien", reason)

    def test_live_backend_detects_next_empty_row_with_single_bulk_request(self) -> None:
        worksheet = self._DummyWorksheet(
            column_values={1: 10.0, 2: 11.0, 3: 12.0, 4: None, 5: 14.0},
            used_row=5,
            used_column=1,
        )
        settings = ExcelSettings(path="dummy.xlsx", sheet_name="Produktion", column="A", start_row=1, mode="live")

        column, row = LIVE_BACKEND._detect_current_cell_fast(worksheet, settings)

        self.assertEqual((column, row), ("A", 4))
        self.assertEqual(worksheet.range_calls, ["A1:A5"])

    def test_live_backend_detects_next_empty_column_with_single_bulk_request(self) -> None:
        worksheet = self._DummyWorksheet(
            row_values={
                column_name_to_index("B"): 10.0,
                column_name_to_index("C"): 11.0,
                column_name_to_index("D"): None,
                column_name_to_index("E"): 14.0,
            },
            used_row=4,
            used_column=column_name_to_index("E"),
        )
        settings = ExcelSettings(
            path="dummy.xlsx",
            sheet_name="Produktion",
            column="B",
            start_row=4,
            direction=FLOW_RIGHT,
            mode="live",
        )

        column, row = LIVE_BACKEND._detect_current_cell_fast(worksheet, settings)

        self.assertEqual((column, row), ("D", 4))
        self.assertEqual(worksheet.range_calls, ["B4:E4"])


if __name__ == "__main__":
    unittest.main()
